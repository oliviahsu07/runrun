import streamlit as st
import io
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOPICS = ["兆豐銀行", "數位帳戶", "信用卡", "流動性風險", "分行"]

HEADER_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NEG_FILL = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
ALT_FILL = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def parse_excel(file_buffer):
    wb = load_workbook(file_buffer, read_only=True, data_only=True)
    counts = {topic: {"正面": 0, "負面": 0, "中立": 0} for topic in TOPICS}
    negatives = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        current_topic = None
        col_map = {}

        for row in rows:
            if not row or row[0] is None:
                continue
            if str(row[0]).strip() == "話題類型":
                current_topic = None
                for topic in TOPICS:
                    for cell in row:
                        if cell and topic in str(cell):
                            current_topic = topic
                            break
                    if current_topic:
                        break
                col_map = {}
                for i, cell in enumerate(row):
                    if not cell:
                        continue
                    name = str(cell).strip()
                    if "話題類型" in name: col_map["話題類型"] = i
                    elif "日期" in name: col_map["日期"] = i
                    elif "標題" in name: col_map["標題"] = i
                    elif "原文網址" in name or "連結" in name: col_map["原文網址"] = i
                    elif "頻道" in name: col_map["頻道"] = i
                    elif "回文數" in name: col_map["回文數"] = i
                    elif "輿情內文" in name or "內文摘要" in name: col_map["輿情內文"] = i
                    elif "正負面" in name or "燈號" in name: col_map["正負面"] = i
                continue

            if current_topic and "正負面" in col_map:
                sentiment_raw = row[col_map["正負面"]]
                if sentiment_raw is None:
                    continue
                sentiment = str(sentiment_raw).strip()
                if sentiment in ("正面", "負面", "中立"):
                    counts[current_topic][sentiment] += 1
                    if sentiment == "負面":
                        negatives.append({
                            "話題分類": current_topic,
                            "話題類型": row[col_map.get("話題類型", 0)],
                            "日期": row[col_map.get("日期", 1)],
                            "輿情標題": row[col_map.get("標題", 2)],
                            "原文網址": row[col_map.get("原文網址", 3)],
                            "頻道": row[col_map.get("頻道", 4)],
                            "回文數": row[col_map.get("回文數", 5)],
                            "輿情內文": row[col_map.get("輿情內文", 6)],
                        })
    wb.close()
    return counts, negatives


def write_summary_sheet(ws, total_counts, month):
    ws.title = f"{month:02d}月彙整"
    headers = ["話題分類", "正面聲量", "中立聲量", "負面聲量", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
    ws.column_dimensions["A"].width = 16
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 12
    for row_idx, topic in enumerate(TOPICS, 2):
        data = total_counts[topic]
        values = [topic, data["正面"], data["中立"], data["負面"],
                  data["正面"] + data["中立"] + data["負面"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = CENTER if col > 1 else LEFT
            cell.border = THIN
            if col == 4 and isinstance(val, int) and val > 0:
                cell.fill = NEG_FILL
    total_row = len(TOPICS) + 2
    ws.cell(row=total_row, column=1, value="Total").border = THIN
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = LEFT
    for col in range(2, 6):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col,
                       value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})")
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.border = THIN


def write_negative_sheet(ws, negatives, month):
    ws.title = f"{month:02d}月負面明細"
    headers = ["話題分類", "話題類型", "日期", "輿情標題", "原文網址", "頻道", "回文數", "輿情內文"]
    col_widths = [12, 10, 18, 40, 50, 18, 10, 60]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = w
    for row_idx, neg in enumerate(negatives, 2):
        values = [neg["話題分類"], neg["話題類型"], neg["日期"], neg["輿情標題"],
                  neg["原文網址"], neg["頻道"], neg["回文數"], neg["輿情內文"]]
        fill = NEG_FILL if row_idx % 2 == 0 else ALT_FILL
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = THIN
            cell.alignment = CENTER if col in (1, 2, 6, 7) else LEFT


# ── Streamlit 介面 ────────────────────────────────────────────────────────
st.set_page_config(page_title="兆豐銀行輿情彙整", page_icon="📊")
st.title("📊 兆豐銀行輿情月報彙整")
st.caption("上傳當月所有每日 Excel 檔，自動統計並產出彙整報表")

col1, col2 = st.columns(2)
with col1:
    year = st.number_input("年份", min_value=2020, max_value=2099, value=2026)
with col2:
    month = st.number_input("月份", min_value=1, max_value=12, value=2)

uploaded_files = st.file_uploader(
    "上傳當月所有每日 Excel 檔（可一次多選）",
    type=["xlsx"],
    accept_multiple_files=True
)

if uploaded_files:
    st.info(f"已選擇 {len(uploaded_files)} 個檔案")

    if st.button("開始彙整", type="primary", use_container_width=True):
        all_counts = []
        all_negatives = []

        progress = st.progress(0)
        status = st.empty()

        for i, f in enumerate(uploaded_files):
            status.text(f"解析中：{f.name}")
            buf = io.BytesIO(f.read())
            counts, negatives = parse_excel(buf)
            all_counts.append(counts)
            all_negatives.extend(negatives)
            progress.progress((i + 1) / len(uploaded_files))

        status.text("產出報表中...")

        total = {topic: {"正面": 0, "負面": 0, "中立": 0} for topic in TOPICS}
        for daily in all_counts:
            for topic in TOPICS:
                for s in ("正面", "負面", "中立"):
                    total[topic][s] += daily[topic][s]

        wb_out = Workbook()
        wb_out.remove(wb_out.active)
        write_summary_sheet(wb_out.create_sheet(), total, month)
        write_negative_sheet(wb_out.create_sheet(), all_negatives, month)

        output = io.BytesIO()
        wb_out.save(output)
        output.seek(0)

        progress.empty()
        status.empty()

        st.success("彙整完成！")

        filename = f"兆豐銀行輿情_Y{str(year)[2:]}_{month:02d}月彙整.xlsx"
        st.download_button(
            label="📥 下載彙整 Excel",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
